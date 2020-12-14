from django.shortcuts import render, redirect
from django.http import HttpResponseRedirect
from django.urls import reverse
from .forms import UploadSheetForm
from openpyxl import Workbook
import openpyxl
from io import BytesIO
from query.models import Query
from multiprocessing import Process, Queue, Pool, cpu_count
from multiprocessing import set_start_method, get_context
from background_task import background
from django.utils import timezone
import time


def upload_file(request):
    """
    View for uploading an excel file with urls.
    Once uploaded, the file will be processed asynchronously.
    """
    if request.method == 'POST':
        form = UploadSheetForm(request.POST, request.FILES)
        if form.is_valid():
            process_sheet(request.FILES['sheet'])
            return redirect('/queries/')
    else:
        form = UploadSheetForm()
    return render(request, 'upload.html', {'form': form})


def process_query(url):
    """
    Task for processing a single query
    """
    query = Query(url=url)
    query.save()


# @background(schedule=1)
def process(urls):
    """
    Task for processing a lot of queries via multithreading,
    run delayed but synchronously.
    Each query is run asynchronously.
    """
    time.sleep(2)
    pool = Pool(cpu_count())
    pool.map(process_query, urls)


def process_sheet(file):
    """
    Iterate over the excel file and process each url
    """
    wb = openpyxl.load_workbook(filename=BytesIO(file.read()))
    sheet = wb.active

    # To speed up, use a Pool
    urls = []

    for row in range(1, sheet.max_row+1):
        url = sheet.cell(row=row, column=1).value
        if'.' not in url:
            continue
        urls.append(url)

    # Process queries on all urls
    process(urls)
